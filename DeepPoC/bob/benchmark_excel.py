import os
import openpyxl

from env import *
from cnn_model import *
from image_treatment import *

class excel_writer():
    def __init__(self):
        self.workbook = openpyxl.load_workbook("benchmark.xlsx")
        self.sheet = self.workbook.active
        self.width = 1
        self.line = 1
        self.next_empty_col()

    def next_empty_col(self):
        while (True):
            if (self.sheet.cell(row=1, column=self.width).value == None):
                break
            self.width += 1

    def add_new_benchmark(self, color, contrast, brightness, sharpness):
        if (self.line != 1):
            self.width += 1
        self.line = 1
        self.sheet.cell(row=self.line, column=self.width).value = "{}, {}, {}, {}".format(color, contrast, brightness, sharpness)
        self.sheet.column_dimensions[openpyxl.utils.get_column_letter(self.width)].width = 25
        self.line += 1

    def add_new_line(self, value):
        self.sheet.cell(row=self.line, column=self.width).value = value
        self.line += 1